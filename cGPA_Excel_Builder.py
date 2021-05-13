from typing import List
from typing import TextIO
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Protection, Font

#---------------------------CONSTANTS--------------------------------------
sheet_list = []
sheet_names = []
sheet_row_index_list = []
sheet_titles_list = []
sheet_subtitles_list = []
sheet_heights_list = []
sheet_percent_index_list = []
sheet_percents_list = []
sheet_types_list = []
sheet_topx_list = []

thin_border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
                    
title_font = Font(name='Calibri', size=14, bold=True)
subtitle_font = Font(name='Calibri', size=14)
ex_font = Font(name='Calibri', size=12)
total_sf_font = Font(name='Calibri', size=12, italic=True)

percent_style =  NamedStyle(name="percent_style")
percent_style.font = ex_font
percent_style.alignment = Alignment(horizontal='left')
percent_style.number_format = '0.00%'

im_sorry = """=IF(AND(E2*100<=100,E2*100>=90),4,
IF(AND(E2*100<=89.99,E2*100>=85),4,
IF(AND(E2*100<=84.99,E2*100>=80),3.7,
IF(AND(E2*100<=79.99,E2*100>=77),3.3,
IF(AND(E2*100<=76.99,E2*100>=73),3,
IF(AND(E2*100<=72.99,E2*100>=70),2.7,
IF(AND(E2*100<=69.99,E2*100>=67),2.3,
IF(AND(E2*100<=66.99,E2*100>=63),2,
IF(AND(E2*100<=62.99,E2*100>=60),1.7,
IF(AND(E2*100<=59.99,E2*100>=57),1.3,
IF(AND(E2*100<=56.99,E2*100>=53),1,
IF(AND(E2*100<=52.99,E2*100>=50),0.7,0))))))))))))""".replace('\n','')

#---------------------------------------------------------------------------


#----------------------------------HELPERS----------------------------------
def set_borders(starter_row: int, height: int) -> None:

    for row in range(starter_row, starter_row+height+2):
            for col in range(1, 4):
                current_sheet.cell(column=col, row=row).border = thin_border


def title_cell(cell_ID: str, title: str) -> None:
    cell = current_sheet[cell_ID]
    cell.font = title_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='left')
    cell.value = title


def enter_section(starter_row: int, height: int,
                  title: str, subtitles: List[str],
                  percentages: List[float], add_type: int,
                  topx: int) -> None:
                  
   
    
    title_cell('A'+str(starter_row), title)
    current_sheet['B'+str(starter_row)].font = subtitle_font
    current_sheet['B'+str(starter_row)].alignment = Alignment(horizontal='center')
    current_sheet['B'+str(starter_row)] = "Grade"
    current_sheet['C'+str(starter_row)].font = subtitle_font
    current_sheet['C'+str(starter_row)] =  "% worth"
    
    for i in range(0, height):
        current_sheet['A'+str(starter_row+i+1)].font = ex_font
        current_sheet['A'+str(starter_row+i+1)] = subtitles[i]
        current_sheet['B'+str(starter_row+i+1)].style = percent_style
        current_sheet['B'+str(starter_row+i+1)].alignment = Alignment(horizontal='center')
        current_sheet['C'+str(starter_row+i+1)].style = percent_style
        current_sheet['C'+str(starter_row+i+1)] = percentages[i]
    
    current_sheet['A'+str(starter_row+height+1)].font = total_sf_font
    current_sheet['A'+str(starter_row+height+1)] = "Total (so far)"
    current_sheet['B'+str(starter_row+height+1)].style = percent_style
    current_sheet['B'+str(starter_row+height+1)].alignment = Alignment(horizontal='center')
    current_sheet['C'+str(starter_row+height+1)].style = percent_style
    
    if (add_type == 1):
        main_formula = "=IFERROR(SUM(B"+str(starter_row+height)+")/COUNTA(B"+str(starter_row+height)+"),0)"
        sec_formula = "=C"+str(starter_row+height)+"*COUNTA(B"+str(starter_row+height)+")"
        
    if (add_type == 2):
        b_range = ("B"+str(starter_row+1)+":B"+str(starter_row+height))
        c_range = ("C"+str(starter_row+1)+":C"+str(starter_row+height))
        main_formula = "=IFERROR(SUM("+b_range+")/COUNTA("+b_range+"),0)"
        sec_formula = "=C"+str(starter_row+height)+"*COUNTA("+b_range+")"
    if (add_type == 3):
        b_range = ("B"+str(starter_row+1)+":B"+str(starter_row+height))
        c_range = ("C"+str(starter_row+1)+":C"+str(starter_row+height))
        main_formula = '=IFERROR(SUMPRODUCT('+b_range+','+c_range+')/SUMPRODUCT(--('+b_range+'<>""), '+c_range+'), 0)'
        sec_formula = '=SUMIF('+b_range+', "<>", '+c_range+')'
    if (add_type == 4):
        b_range = ("B"+str(starter_row+1)+":B"+str(starter_row+height))
        c_range = ("C"+str(starter_row+1)+":C"+str(starter_row+height))
        main_formula = "=MAX(IFERROR(SUM("+b_range+")/COUNTA("+b_range+"),0), IFERROR(AVERAGE(LARGE("+b_range+",{"
        main_formula += ",".join(str(int) for int in list(range(1,topx+1)))
        main_formula += "})), IFERROR(SUM("+b_range+")/COUNTA("+b_range+"),0)))"
        sec_formula = "=MIN(C"+str(starter_row+height)+"*"+str(topx)+", C"+str(starter_row+height)+"*COUNTA("+b_range+"))"
    
    current_sheet['B'+str(starter_row+height+1)] = main_formula
    current_sheet['C'+str(starter_row+height+1)] = sec_formula
    
    set_borders(starter_row, height)
    
    sheet_row_index_list[current_index] += (height+3)
    sheet_titles_list[current_index].append(title)
    sheet_subtitles_list[current_index].append(subtitles)
    sheet_heights_list[current_index].append(height)
    sheet_percent_index_list[current_index].append(sheet_row_index_list[current_index]-2)
    sheet_percents_list[current_index].append(percentages)
    sheet_types_list[current_index].append(add_type)
    sheet_topx_list[current_index].append(topx)
    


#---------------------------------------------------------------------------



dest_filename = input("Input the spreadsheet's name: ")

wb = Workbook()
ws1 = wb.active
sheet_list.append(ws1)
sheet_names.append(ws1.title)
sheet_row_index_list.append(1)
sheet_titles_list.append([])
sheet_subtitles_list.append([])
sheet_heights_list.append([])
sheet_percent_index_list.append([])
sheet_percents_list.append([])
sheet_types_list.append([])
sheet_topx_list.append([])

current_sheet = ws1
current_index = 0
 
option_select = -1

while (option_select != '0'):
    print("-------------------------------------------------------------")
    print("Worksheet name: "+dest_filename)
    print("Current sheet: "+current_sheet.title)
    print("Sheets: "+", ".join(sheet_names))
    print("-------------------------------------------------------------")
    print("1. Change worksheet name")
    print("2. Change sheet name")
    print("3. Change current sheet")
    print("4. Add new sheet")
    print("5. Enter section with same entries, repeated %")
    print("6. Enter section with same entries, different %")
    print("7. Enter section with same entries, repeated % [top x of y]")
    print("8. Enter section with big boye/lone entry")
    print("9. Save worksheet")
    print("10. Save current workplace")
    print("11. Open workplace settings")
    print("12. Delete entry in current sheet")
    print("0. Exit")
    option_select = input("Input an option: ")
    
    if (option_select == '1'):
        dest_filename = input("Input new worksheet name: ")
        
        
    if (option_select == '2'):
        current_sheet.title = input("Input new sheet name: ")
        sheet_names[current_index] = current_sheet.title
        
        
    if (option_select == '3'):
        temp = input("Input sheet to change to: ")
        if (temp in sheet_names):
            current_sheet = sheet_list[sheet_names.index(temp)]
            current_index = sheet_names.index(temp)
            print("Changed sheet to "+temp)
        else:
            print("Unable to find sheet in current sheet list")
            
            
    if (option_select == '4'):
        temp = input("Input new sheet's name: ")
        sheet_list.append(wb.create_sheet(temp))
        sheet_names.append(sheet_list[-1].title)
        sheet_row_index_list.append(1)
        sheet_titles_list.append([])
        sheet_subtitles_list.append([])
        sheet_heights_list.append([])
        sheet_percent_index_list.append([])
        sheet_percents_list.append([])
        sheet_types_list.append([])
        sheet_topx_list.append([])
        print("Created sheet '"+temp+"', changing to new sheet...")
        current_sheet = sheet_list[sheet_names.index(temp)]
        current_index = sheet_names.index(temp)
        
        
    if (option_select == '5'):
        title = input("Input title of entry: ")
        subtitle = input("Input subtitles of entry: ")
        number = int(input("Input number of subtitles: "))
        worth = float(input("Input worth in % (decimals allowed): "))
        
        sub_list = []
        for i in range(0, number):
            sub_list.append(subtitle+" "+str(i+1))
        
        percent = []
        for i in range(0, number):
            percent.append((worth/number)*0.01)
        
        enter_section(sheet_row_index_list[current_index], number, 
                      title, sub_list, percent,2,0)
        
        
        
    if (option_select == '6'):
        title = input("Input title of entry: ")
        subtitle = input("Input subtitles of entry: ")
        number = int(input("Input number of subtitles: "))
        
        sub_list = []
        percent = []
        for i in range(0, number):
            sub_list.append(subtitle+" "+str(i+1))
            percent.append(float(input("Input worth in % for "+subtitle+" "+str(i+1)+": "))*0.01)
        
        enter_section(sheet_row_index_list[current_index], number, 
                      title, sub_list, percent,3,0)
        
        
    if (option_select == '7'):
        title = input("Input title of entry: ")
        subtitle = input("Input subtitles of entry: ")
        number = int(input("Input number of subtitles: "))
        worth = float(input("Input worth in % (decimals allowed): "))
        topx = int(input("Input top x of "+str(number)+" to take into account: "))
        
        sub_list = []
        for i in range(0, number):
            sub_list.append(subtitle+" "+str(i+1))
        
        percent = []
        for i in range(0, number):
            percent.append((worth/topx)*0.01)
        
        enter_section(sheet_row_index_list[current_index], number, 
                      title, sub_list, percent,4,topx)
        
    if (option_select == '8'):
        title = input("Input title of entry: ")
        subtitle = input("Input subtitle of entry: ")
        worth = float(input("Input worth in % (decimals allowed): "))
        enter_section(sheet_row_index_list[current_index], 1, title,
                     [subtitle], [worth*0.01],1,0)
        
        
    if (option_select == '9'):
        for sheet in sheet_list:
            current_sheet = sheet
            current_index = sheet_list.index(current_sheet)
            sheet.column_dimensions['A'].width = 37
            sheet.column_dimensions['B'].width = 60
            sheet.column_dimensions['C'].width = 19
            sheet.column_dimensions['E'].width = 40
            sheet.column_dimensions['F'].width = 22
            title_cell("E1", "Current Mark")
            title_cell("F1", "Course Completion")
            title_cell("E4", "Mark Needed")
            title_cell("E7", "Remaining Mark Avg. to Achieve")
            title_cell("E10", "GPA")
            title_cell("F10", "Desired GPA")
            title_cell("E14", "Mark Override")
            sheet["E2"].style = percent_style
            sheet["F2"].style = percent_style
            sheet["E5"].style = percent_style
            sheet["E8"].style = percent_style
            sheet["E15"].style = percent_style
            sheet["E8"] = "=(E5-(E2*F2))/(1-F2)"
            sheet["E11"] = im_sorry
            sheet["E11"].font = ex_font
            sheet["E11"].alignment = Alignment(horizontal='left')
            sheet["F11"] = im_sorry.replace("E2", "E5")
            sheet["F11"].font = ex_font
            sheet["F11"].alignment = Alignment(horizontal='left')
            
            mark_str = "=IF(ISBLANK(E15), (SUM("
            for i in sheet_percent_index_list[current_index]:
                mark_str += "(B"+str(i)+"*C"+str(i)+"),"
            mark_str += ")/SUM("
            for i in sheet_percent_index_list[current_index]:
                mark_str += "C"+str(i)+","
            mark_str += ")), E15)"
            sheet["E2"] = mark_str
            
            course_str = "=SUM("
            for i in sheet_percent_index_list[current_index]:
                course_str += "C"+str(i)+","
            course_str += ")"
            sheet["F2"] = course_str
        
        cGPA = wb.create_sheet("cGPA")
        current_index = sheet_list.index(current_sheet)
        temp_s = current_sheet
        current_sheet = cGPA
        cGPA.column_dimensions['D'].width = 21
        title_cell("D4", "Current Mark")
        title_cell("D7", "cGPA")
        title_cell("D10", "Desired cGPA")
        cGPA["D5"].style = percent_style
        cGPA["D8"].font = ex_font
        cGPA["D8"].alignment = Alignment(horizontal='left')
        cGPA["D11"].font = ex_font
        cGPA["D11"].alignment = Alignment(horizontal='left')
        
        cGPA_current = "=SUM("
        for sheet in sheet_list:
            cGPA_current += sheet.title+"!E2,"
        cGPA_current += ")/"+str(len(sheet_list))
        
        cGPA["D5"] = cGPA_current
        cGPA["D8"] = cGPA_current.replace("!E2,", "!E11,")
        cGPA["D11"] = cGPA_current.replace("!E2,", "!F11,")
            
        wb.save(dest_filename+".xlsx")
        print("Saved!")
        wb.remove(cGPA)
        current_sheet = temp_s
        
        
    if (option_select == '10'):
        filename = input("Select name for saving settings (no extension): ")
        file = open(filename+".txt", 'w')
        file.write(dest_filename+'\n')
        file.write(str(len(sheet_names))+'\n\n\n')
        
        for i in range(0, len(sheet_names)):
            file.write(sheet_names[i]+'\n')
            file.write(str(len(sheet_titles_list[i]))+'\n')
            for x in range(0, len(sheet_titles_list[i])):
                file.write(sheet_titles_list[i][x]+'\n')
                file.write(str(len(sheet_subtitles_list[i][x]))+'\n')
                for a in range(0, len(sheet_subtitles_list[i][x])):
                    file.write(sheet_subtitles_list[i][x][a]+'\n')
                    file.write(str(sheet_percents_list[i][x][a])+'\n')
                file.write(str(sheet_heights_list[i][x])+'\n')
                file.write(str(sheet_types_list[i][x])+'\n')
                file.write(str(sheet_topx_list[i][x])+'\n')
            file.write('\n')
        
        file.close()
    
    
    if (option_select == '11'):
        filename = input("Input filename (extension included): ")
        file = open(filename, 'r')
        
        for sheet in sheet_list:
            wb.remove(sheet)
        
        sheet_list = []
        sheet_names = []
        sheet_row_index_list = []
        sheet_titles_list = []
        sheet_subtitles_list = []
        sheet_heights_list = []
        sheet_percent_index_list = []
        sheet_percents_list = []
        sheet_types_list = []
        sheet_topx_list = []
        
        
        dest_filename = file.readline()[:-1]
        sheet_no = int(file.readline())
        file.readline()
        file.readline()
        for i in range(0, sheet_no):
            
            
            sheet_list.append(wb.create_sheet(file.readline()[:-1]))
            current_index = i
            sheet_names.append(sheet_list[-1].title)
            current_sheet = sheet_list[i]
            sheet_row_index_list.append(1)
            sheet_titles_list.append([])
            sheet_subtitles_list.append([])
            sheet_heights_list.append([])
            sheet_percent_index_list.append([])
            sheet_percents_list.append([])
            sheet_types_list.append([])
            sheet_topx_list.append([])
            
            for x in range(0, int(file.readline())):
                title = file.readline()[:-1]
                subtitles = []
                percentages = []
                for a in range(0, int(file.readline())):
                    subtitles.append(file.readline()[:-1])
                    percentages.append(float(file.readline()))
                height = int(file.readline())
                add_type = int(file.readline())
                topx = int(file.readline())
                enter_section(sheet_row_index_list[current_index],
                              height, title, subtitles, percentages, add_type,topx)
            file.readline()
            
    
    if (option_select == '12'):
        if (len(sheet_titles_list[current_index]) == 0):
            print ("There are no entries to delete.")
        else:
            print("Titles: "+", ".join(sheet_titles_list[current_index]))
            delete = input("Input the title of the entry to delete: ")
            if (delete in sheet_titles_list[current_index]):
                del_index = sheet_titles_list[current_index].index(delete)
                print(sheet_percent_index_list[current_index])
                current_sheet.delete_rows(sheet_percent_index_list[current_index][del_index]-
                                          sheet_heights_list[current_index][del_index]-1,
                                          sheet_percent_index_list[current_index][del_index]+1)
                sheet_row_index_list[current_index] -= (sheet_heights_list[current_index][del_index]+3)
                sheet_subtitles_list[current_index].pop(del_index)
                for i in range(del_index, len(sheet_percent_index_list[current_index])):
                    sheet_percent_index_list[current_index][i] -= sheet_heights_list[current_index][del_index]+3
                sheet_percent_index_list[current_index].pop(del_index)
                print(sheet_percent_index_list[current_index])
                sheet_heights_list[current_index].pop(del_index)
                sheet_types_list[current_index].pop(del_index)
                sheet_topx_list[current_index].pop(del_index)
                print("Deleting "+delete+"...")
                sheet_titles_list[current_index].pop(del_index)
                
            else:
                print("Title entry not found.")
            
        